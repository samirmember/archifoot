#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build:
  - matches_md_structured.xlsx (MD -> table normalisée)
  - en_matches_complet.xlsx (en_matches.xlsx complété + colonnes ajoutées + couleurs)

Contraintes clés:
- 1 cellule = 1 info
- Jointure par N° du match
- Dates ISO YYYY-MM-DD
- Buts: "Prénom Nom 61'" + optionnel " (penalty)" ou " (own_goal)" uniquement
- Remplacements: parenthèse = entrant ; sortant = joueur précédent ; minute si dispo -> côté sortant
- Couleurs:
    VERT   : vide -> rempli depuis MD
    ORANGE : déjà rempli -> amélioré (format, précision, nom complet si inclus)
"""

import argparse
import re
import unicodedata
from datetime import datetime
from typing import List, Dict, Tuple, Optional

import pandas as pd
from dateutil import parser as dtparser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ------------------------------
# Utils normalisation
# ------------------------------

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")   # vert clair
ORANGE_FILL = PatternFill("solid", fgColor="FFEB9C")  # orange clair

def strip_md_decor(s: str) -> str:
    s = s.strip()
    # retire markup courant markdown
    s = re.sub(r"^[#\s>*`_]+", "", s)
    s = s.replace("**", "").replace("__", "")
    return s.strip()

def nfd(s: str) -> str:
    return unicodedata.normalize("NFD", s)

def norm_key(s: str) -> str:
    if s is None:
        return ""
    s = s.strip()
    s = "".join(ch for ch in nfd(s) if unicodedata.category(ch) != "Mn")
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("’", "'")
    return s

def is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")

def iso_date_from_text(s: str) -> Optional[str]:
    """
    Détecte dd.mm.yyyy, d.m.yy, etc. et renvoie YYYY-MM-DD.
    """
    s = s.replace("’", "'")
    m = re.search(r"(\d{1,2})[./](\d{1,2})[./](\d{2,4})", s)
    if not m:
        return None
    d, mth, y = m.group(1), m.group(2), m.group(3)
    if len(y) == 2:
        # heuristique: 63 => 1963, 05 => 2005 (adapter si besoin)
        y = "19" + y if int(y) >= 40 else "20" + y
    try:
        dt = datetime(int(y), int(mth), int(d))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None

def split_outside_parens(s: str, sep: str = ",") -> List[str]:
    """
    Split sur virgules hors parenthèses.
    """
    out, buf, depth = [], [], 0
    for ch in s:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        if ch == sep and depth == 0:
            out.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        out.append("".join(buf).strip())
    return [x for x in out if x]

def clean_player_token(tok: str) -> Tuple[str, bool]:
    """
    Extrait marqueur capitaine (c/cap/©/C).
    Retourne (nom_sans_marqueur, is_captain).
    """
    t = tok.strip()
    is_cap = False
    if re.search(r"\((c|cap|C)\)", t, flags=re.I) or "©" in t:
        is_cap = True
    t = t.replace("©", "")
    t = re.sub(r"\((c|cap|C)\)", "", t, flags=re.I).strip()
    t = re.sub(r"\s+", " ", t)
    return t, is_cap


# ------------------------------
# Parsing MD : blocs matchs
# ------------------------------

HEADER_RE = re.compile(r"^\(?\s*(\d+)\s*\)\s*(.+)$")
SCORE_RE = re.compile(r"^(.*?)\s+(\d+)\s+(.+?)\s+(\d+)\s*$", re.U)

def iter_md_matches(md_text: str) -> List[Dict]:
    lines = [strip_md_decor(l) for l in md_text.splitlines()]
    matches = []
    cur = None

    def flush():
        nonlocal cur
        if cur:
            matches.append(cur)
            cur = None

    for raw in lines:
        line = raw.strip()
        if not line:
            continue

        m = HEADER_RE.match(line)
        if m:
            num = m.group(1)
            rest = m.group(2).strip()
            sm = SCORE_RE.match(rest)
            if sm:
                flush()
                cur = {
                    "match_no": int(num),
                    "team_a": sm.group(1).strip(),
                    "score_a": int(sm.group(2)),
                    "team_b": sm.group(3).strip(),
                    "score_b": int(sm.group(4)),
                    "raw_lines": []
                }
                continue

        if cur:
            cur["raw_lines"].append(line)

    flush()
    return matches


# ------------------------------
# Parsing détails match
# ------------------------------

def parse_referee(text: str) -> Tuple[str, str]:
    """
    Retourne (arbitre_principal, nationalite_arbitre)
    """
    t = text.replace("’", "'")
    # "Arbitrage de M. Nom (Nat)."
    m = re.search(r"Arbitrage\s*(?:de)?\s*(?:M\.|Mr\.|Monsieur)?\s*([A-Za-zÀ-ÖØ-öø-ÿ\-\s'.]+?)\s*(?:\(([^)]+)\))?(?:[.,;-]|$)", t, flags=re.I)
    if m:
        name = m.group(1).strip()
        nat = (m.group(2) or "").strip()
        name = re.sub(r"^(M\.|Mr\.)\s*", "", name).strip()
        return name, nat

    # fallback: "Taylor (Angleterre)." isolé
    m2 = re.search(r"\b([A-Za-zÀ-ÖØ-öø-ÿ\-\s'.]{2,})\s*\(([^)]+)\)", t)
    if m2 and "buts" not in norm_key(t) and "entra" not in norm_key(t):
        return m2.group(1).strip(), m2.group(2).strip()

    return "", ""

def parse_location_and_competition(line: str) -> Tuple[str, str, str, str]:
    """
    Sur la ligne date/lieu/compétition:
      date_iso, stade, ville, competition_edition_etape_raw
    """
    date_iso = iso_date_from_text(line) or ""
    # coupe autour de la date
    # exemple: "06.01.1963, Stade 20 Août (Alger). Match amical. Arbitrage..."
    after = line
    if date_iso:
        # retire le préfixe date
        after = re.sub(r"^\s*\d{1,2}[./]\d{1,2}[./]\d{2,4}\s*[:,]?\s*", "", after).strip()

    stade, ville = "", ""
    comp_raw = ""

    # première phrase jusqu'au premier point = lieu
    parts = after.split(".", 1)
    loc = parts[0].strip(" ,")
    rest = parts[1].strip() if len(parts) > 1 else ""

    # stade (ville) ?
    m = re.search(r"^(.*?)\s*\(([^)]+)\)\s*$", loc)
    if m:
        stade = m.group(1).strip(" ,")
        ville = m.group(2).strip(" ,")
    else:
        # pas de parenthèse => ville brute
        # si contient "Stade", garder en stade
        if re.search(r"\bstade\b", loc, flags=re.I):
            stade = loc
        else:
            ville = loc

    # compétition brute = avant "Arbitrage" ou "Buts"/"But"
    rest2 = rest
    cut = re.split(r"\b(Arbitrage|Buts?\s*:|but\s*:|Cartons?)\b", rest2, flags=re.I)[0]
    comp_raw = cut.strip(" .-")

    return date_iso, stade, ville, comp_raw

def parse_coaches(line: str) -> Tuple[str, str, str, str]:
    """
    Retourne: entraineur_principal, assistant1, assistant2, notes
    """
    t = line
    t = t.replace("Entraîneurs", "Entraineurs").replace("Entraîneur", "Entraineur")
    m = re.search(r"\bEntraineurs?\s*:\s*(.+)$", t, flags=re.I)
    if not m:
        return "", "", "", ""
    payload = m.group(1).strip().strip(".")
    # séparateurs: ".", ",", " et "
    # on remplace " et " par ","
    payload2 = re.sub(r"\s+et\s+", ",", payload, flags=re.I)
    payload2 = payload2.replace(";", ",")
    # split sur "." puis ","
    chunks = []
    for p in payload2.split("."):
        p = p.strip()
        if p:
            chunks.extend([x.strip() for x in p.split(",") if x.strip()])

    # dédoublonnage conservant ordre
    seen = set()
    names = []
    for c in chunks:
        k = norm_key(c)
        if k and k not in seen:
            names.append(c)
            seen.add(k)

    ent, a1, a2, note = "", "", "", ""
    if len(names) == 1:
        ent = names[0]
    elif len(names) == 2:
        ent, a1 = names[0], names[1]
        note = "2 coachs listés (principal+assistant par heuristique)"
    elif len(names) >= 3:
        ent, a1, a2 = names[0], names[1], names[2]
        if len(names) > 3:
            note = f"{len(names)} coachs listés (seuls 3 champs coach utilisés, surplus en note)"
            note += " | autres=" + " / ".join(names[3:])
    return ent, a1, a2, note

def parse_players_line(line: str) -> Dict:
    """
    Parse "Equipe : joueurs..., X (Y, 61'), (Z), ..." + capitaine
    Retourne:
      team, starters[11], entrants[<=9], subs_out[<=10], captain
    """
    # "Algérie : ..."
    m = re.match(r"^([^:]+)\s*:\s*(.+)$", line)
    if not m:
        return {}
    team = m.group(1).strip()
    payload = m.group(2).strip().strip(".")
    tokens = split_outside_parens(payload, ",")
    starters = []
    entrants = []
    subs_out = []  # "Nom Prenom 61'" (minute optionnelle)
    captain = ""

    last_starter = ""
    for tok in tokens:
        tok = tok.strip().strip(".")
        if not tok:
            continue

        # parenthèse seule => entrant
        if re.match(r"^\(.+\)$", tok):
            inn = tok.strip("() ").strip()
            inn, is_cap = clean_player_token(inn)
            if last_starter:
                out = last_starter
                subs_out.append(out)  # minute inconnue
                entrants.append(inn)
            continue

        # forme: OUT (IN, 61') ou OUT (IN)
        msub = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", tok)
        if msub:
            out_raw = msub.group(1).strip()
            inn_raw = msub.group(2).strip()
            out_name, out_cap = clean_player_token(out_raw)

            # IN peut contenir ", 61'"
            inn_name = inn_raw
            minute = ""
            mm = re.search(r"(.*?)(?:,|\s)\s*(\d{1,3})\s*[’']?\s*$", inn_raw)
            if mm and re.search(r"\d{1,3}", mm.group(2)):
                inn_name = mm.group(1).strip()
                minute = mm.group(2).strip()
            inn_name, inn_cap = clean_player_token(inn_name)

            # starters = OUT
            if len(starters) < 20:
                starters.append(out_name)
            last_starter = out_name
            if out_cap and not captain:
                captain = out_name

            # substitution
            out_val = out_name
            if minute:
                out_val = f"{out_name} {minute}'"
            subs_out.append(out_val)
            entrants.append(inn_name)
            continue

        # joueur simple
        name, is_cap = clean_player_token(tok)
        if len(starters) < 20:
            starters.append(name)
        last_starter = name
        if is_cap and not captain:
            captain = name

    # starters réels = les 11 premiers
    starters11 = starters[:11]
    # entrants -> banc j12.. (max 9)
    entrants9 = entrants[:9]
    subs_out10 = subs_out[:10]

    # compléter joueurs jusqu’à 20 (j12..j20 = entrants puis vide)
    players20 = starters11 + entrants9
    players20 += [""] * (20 - len(players20))

    return {
        "team": team,
        "players20": players20,
        "subs_out10": subs_out10 + [""] * (10 - len(subs_out10)),
        "captain": captain
    }

def normalize_goal_item(name: str, minute: str, flag: str) -> str:
    """
    flag in {"", "penalty", "own_goal"}
    """
    base = f"{name} {minute}'"
    if flag == "penalty":
        return base + " (penalty)"
    if flag == "own_goal":
        return base + " (own_goal)"
    return base

def parse_goals_segment(seg: str) -> List[str]:
    """
    Parse une séquence de buts (sans 'pour XXX'), retourne liste normalisée.
    Heuristiques:
      - duplique minutes "X 40' et 85'" => 2 entrées
      - si 'PEN' présent et plusieurs minutes => visible sur la dernière minute
    """
    s = seg.replace("’", "'")
    # normalise minute entre parenthèses: (65') => 65'
    s = re.sub(r"\((\d{1,3})\s*'\)", r"\1'", s)

    # marqueurs
    s = re.sub(r"\b(pen\.?|penalty|sur\s+p[ée]nalty)\b", " PEN ", s, flags=re.I)
    s = re.sub(r"\b(csc|contre\s+son\s+camp)\b", " OG ", s, flags=re.I)

    # split sur virgules d'abord
    toks = [t.strip() for t in split_outside_parens(s, ",") if t.strip()]
    out = []
    last_name = ""

    for t in toks:
        # split complémentaire " et " (mais garder minutes nues)
        parts = [p.strip() for p in re.split(r"\s+\bet\b\s+", t, flags=re.I) if p.strip()]
        for part in parts:
            mins = re.findall(r"(\d{1,3})\s*'", part)
            has_pen = "PEN" in part
            has_og = "OG" in part
            # retire minutes et marqueurs pour isoler le nom
            name = re.sub(r"\bPEN\b|\bOG\b", "", part).strip()
            name = re.sub(r"\d{1,3}\s*'", "", name).strip(" .-;:")
            name = re.sub(r"\s+", " ", name).strip()

            if name:
                last_name = name
            else:
                name = last_name

            if not name or not mins:
                continue

            # flag sur la dernière minute si multiple
            for i, mn in enumerate(mins):
                flag = ""
                if has_pen and i == len(mins) - 1:
                    flag = "penalty"
                if has_og and i == len(mins) - 1:
                    flag = "own_goal"
                out.append(normalize_goal_item(name, mn, flag))

    return out

def parse_goals(text: str, team_a: str, team_b: str, score_a: int, score_b: int) -> Tuple[List[str], List[str], str]:
    """
    Retourne (goals_a, goals_b, notes)
    """
    t = text.replace("’", "'")
    m = re.search(r"\bButs?\s*:\s*(.+)$", t, flags=re.I)
    if not m:
        # parfois "but :" / "But :"
        m = re.search(r"\bbut\s*:\s*(.+)$", t, flags=re.I)
    if not m:
        return [], [], ""

    payload = m.group(1).strip().strip(".")
    notes = ""

    # découpe par "pour XXX"
    # capture: "<segment>, pour <team>"
    pat = re.compile(r"(.*?)\s*,?\s*pour\s+(?:l['’]|la|le)?\s*([^,.;]+)\s*(?:,|;|\.|$)", flags=re.I)
    pos = 0
    chunks = []
    for mm in pat.finditer(payload):
        seg = mm.group(1).strip(" ,")
        team = mm.group(2).strip()
        chunks.append((seg, team))
        pos = mm.end()

    tail = payload[pos:].strip(" ,")
    # Cas avec "pour" => chaque seg attribué explicitement
    goals_a, goals_b = [], []
    if chunks:
        for seg, team in chunks:
            gl = parse_goals_segment(seg)
            if norm_key(team) in norm_key(team_a):
                goals_a.extend(gl)
            elif norm_key(team) in norm_key(team_b):
                goals_b.extend(gl)
            else:
                # équipe non reconnue -> note
                notes += f"TeamInGoalsUnmatched={team} | "
        # tail: s'il reste et qu'il n'a pas de 'pour', ambigu
        if tail:
            notes += "TailGoalsAmbiguous | "
    else:
        # Pas de 'pour' -> heuristique si total correspond au score
        flat = parse_goals_segment(payload)
        total = score_a + score_b
        if total > 0 and len(flat) == total:
            # heuristique: premiers score_a pour team_a, reste pour team_b
            goals_a = flat[:score_a]
            goals_b = flat[score_a:]
            notes += "GoalsAssignedByScoreHeuristic(no 'pour') | "
        else:
            notes += "GoalsAmbiguous(no 'pour') | "

    return goals_a[:15], goals_b[:15], notes.strip(" |")


def parse_one_match(m: Dict) -> Dict:
    team_a, team_b = m["team_a"], m["team_b"]
    score_a, score_b = m["score_a"], m["score_b"]

    # ligne "date/lieu/compétition/arbitrage/buts" = première contenant une date
    date_line = ""
    for l in m["raw_lines"]:
        if iso_date_from_text(l):
            date_line = l
            break

    date_iso, stade, ville, comp_raw = ("", "", "", "")
    if date_line:
        date_iso, stade, ville, comp_raw = parse_location_and_competition(date_line)

    # arbitre et buts: chercher dans toutes les lignes
    all_text = " ".join(m["raw_lines"])
    arb, arb_nat = parse_referee(all_text)
    goals_a, goals_b, goals_notes = parse_goals(all_text, team_a, team_b, score_a, score_b)

    # joueurs & coachs
    dz = {}
    adv = {}
    coach_dz = ("", "", "", "")
    coach_adv = ("", "", "", "")
    notes = []

    # contexte: l'ordre des lignes
    ctx = None
    for l in m["raw_lines"]:
        if re.match(r"^Alg[ée]rie\s*:", l, flags=re.I):
            ctx = "dz"
            dz = parse_players_line(l)
            continue
        if re.match(rf"^{re.escape(team_a)}\s*:", l) or re.match(rf"^{re.escape(team_b)}\s*:", l):
            # ligne équipe adverse si ce n'est pas Algérie
            if not re.match(r"^Alg[ée]rie\s*:", l, flags=re.I):
                ctx = "adv"
                adv = parse_players_line(l)
            continue

        if re.search(r"\bEntraineurs?\s*:", l, flags=re.I):
            ent, a1, a2, n = parse_coaches(l)
            if ctx == "dz":
                coach_dz = (ent, a1, a2, n)
            elif ctx == "adv":
                coach_adv = (ent, a1, a2, n)
            else:
                # ambigu: pas de contexte
                notes.append("CoachLineWithoutTeamContext")
            continue

        # TODO cartes: si tu as des lignes "Cartons ..." dans le MD, parser ici

    # Identifie côté Algérie vs adverse dans header
    a_is_dz = norm_key(team_a) == norm_key("Algérie") or "alger" in norm_key(team_a)
    b_is_dz = norm_key(team_b) == norm_key("Algérie") or "alger" in norm_key(team_b)

    # défaut: on attend toujours Algérie présente
    # mais on force côté dz/adv selon l'équipe réellement "Algérie"
    if not dz and (a_is_dz or b_is_dz):
        # si la ligne "Algérie :" manque, on laisse vide + note
        notes.append("MissingAlgeriaLineInMD")

    # goals: répartir selon équipe A/B puis remapper dz/adv
    dz_goals = goals_a if a_is_dz else (goals_b if b_is_dz else [])
    adv_goals = goals_b if a_is_dz else (goals_a if b_is_dz else [])

    out = {
        "match_no": m["match_no"],
        "date": date_iso,
        "team_a": team_a,
        "team_b": team_b,
        "score_a": score_a,
        "score_b": score_b,
        "stade": stade,
        "ville": ville,
        "competition_edition_etape": comp_raw,
        "arbitre_principal": arb,
        "nationalite_arbitre": arb_nat,
        "notes_global": " | ".join([x for x in [goals_notes] + notes if x]).strip(" |"),
    }

    # DZ players/subs/captain
    dz_players = dz.get("players20", [""] * 20)
    dz_subs = dz.get("subs_out10", [""] * 10)
    out["capitaine_dz"] = dz.get("captain", "")
    for i in range(20):
        out[f"dz_j{i+1}"] = dz_players[i] if i < len(dz_players) else ""
    for i in range(10):
        out[f"dz_chang{i+1}"] = dz_subs[i] if i < len(dz_subs) else ""
    for i in range(15):
        out[f"dz_but{i+1}"] = dz_goals[i] if i < len(dz_goals) else ""
    out["entraineur_dz"] = coach_dz[0]
    out["assistant1_dz"] = coach_dz[1]
    out["assistant2_dz"] = coach_dz[2]
    if coach_dz[3]:
        out["notes_global"] = (out["notes_global"] + " | " + coach_dz[3]).strip(" |")

    # ADV players/subs/captain
    adv_players = adv.get("players20", [""] * 20)
    adv_subs = adv.get("subs_out10", [""] * 10)
    out["capitaine_adv"] = adv.get("captain", "")
    for i in range(20):
        out[f"adv_j{i+1}"] = adv_players[i] if i < len(adv_players) else ""
    for i in range(10):
        out[f"adv_chang{i+1}"] = adv_subs[i] if i < len(adv_subs) else ""
    for i in range(15):
        out[f"adv_but{i+1}"] = adv_goals[i] if i < len(adv_goals) else ""
    out["entraineur_adv"] = coach_adv[0]
    out["assistant1_adv"] = coach_adv[1]
    out["assistant2_adv"] = coach_adv[2]
    if coach_adv[3]:
        out["notes_global"] = (out["notes_global"] + " | " + coach_adv[3]).strip(" |")

    # colonnes cartons (optionnelles, vides par défaut)
    for i in range(5):
        out[f"carton_rouge_{i+1}"] = ""
        out[f"carton_jaune_{i+1}"] = ""
        out[f"adv_carton_rouge_{i+1}"] = ""
        out[f"adv_carton_jaune_{i+1}"] = ""

    # arbitres assistants (optionnels)
    out["arbitre_assistant1"] = ""
    out["nationalite_arbitre_assistant1"] = ""
    out["arbitre_assistant2"] = ""
    out["nationalite_arbitre_assistant2"] = ""
    out["quatrieme_arbitre"] = ""
    out["nationalite_quatrieme_arbitre"] = ""

    return out


# ------------------------------
# XLSX complétion + coloration
# ------------------------------

def detect_header_row(ws, max_scan=10) -> int:
    """
    Heuristique: trouve la première ligne qui contient "match" et "N°" (ou "No").
    """
    for r in range(1, max_scan + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        nk = " ".join(norm_key(str(v)) for v in vals if v is not None)
        if "match" in nk and ("n" in nk or "no" in nk):
            return r
    return 1

def build_colmap(ws, header_row: int) -> Dict[str, int]:
    colmap = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        k = norm_key(str(v))
        if k:
            colmap[k] = c
    return colmap

def ensure_columns(ws, header_row: int, wanted_cols: List[str]) -> Dict[str, int]:
    """
    Ajoute les colonnes manquantes en fin de feuille.
    wanted_cols : liste de noms EXACTS souhaités (tels qu'affichés en header).
    Retourne map: nom_exact -> index_col
    """
    existing = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            existing[v.strip()] = c

    col_index = {name: existing[name] for name in wanted_cols if name in existing}
    for name in wanted_cols:
        if name not in col_index:
            ws.cell(header_row, ws.max_column + 1, value=name)
            col_index[name] = ws.max_column
    return col_index

def find_matchno_col(ws, header_row: int) -> int:
    # cherche "N° du match" (avec variantes)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        k = norm_key(str(v))
        if "match" in k and ("n" in k or "no" in k or "numero" in k):
            return c
    raise RuntimeError("Colonne 'N° du match' introuvable dans en_matches.xlsx")

def is_improvement(old: str, new: str) -> bool:
    """
    Heuristiques d'amélioration (ORANGE):
      - new contient old + plus d'info (nom complet, type but, etc.)
      - format minute "61'" ou suffixes (penalty/own_goal)
      - date ISO vs autre format
    """
    o, n = (old or ""), (new or "")
    ko, kn = norm_key(o), norm_key(n)
    if ko == kn:
        return False
    if ko and ko in kn and len(n) > len(o):
        return True
    # ajout (penalty)/(own_goal)
    if ("(penalty)" in n or "(own_goal)" in n) and "(penalty)" not in o and "(own_goal)" not in o:
        return True
    # ajout apostrophe minute
    if re.search(r"\d{1,3}'", n) and not re.search(r"\d{1,3}'", o):
        return True
    # date ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}$", n) and not re.match(r"^\d{4}-\d{2}-\d{2}$", o):
        return True
    return False

def apply_cell(ws, row: int, col: int, new_val: str, stats: Dict[str, int]):
    cell = ws.cell(row, col)
    old_val = cell.value
    old_str = "" if old_val is None else str(old_val).strip()
    new_str = "" if new_val is None else str(new_val).strip()

    if not new_str:
        return

    if is_blank(old_val):
        cell.value = new_str
        cell.fill = GREEN_FILL
        stats["filled_green"] += 1
    else:
        if norm_key(old_str) == norm_key(new_str):
            return
        if is_improvement(old_str, new_str):
            cell.value = new_str
            cell.fill = ORANGE_FILL
            stats["improved_orange"] += 1
        else:
            # conflit: on conserve l'ancien, on note ailleurs
            stats["conflicts"] += 1

def update_xlsx(en_xlsx_path: str, out_path: str, df_md: pd.DataFrame) -> Dict[str, int]:
    wb = load_workbook(en_xlsx_path)
    ws = wb.active
    header_row = detect_header_row(ws)
    matchno_col = find_matchno_col(ws, header_row)

    # index match_no -> row
    idx = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, matchno_col).value
        if v is None:
            continue
        try:
            idx[int(str(v).strip())] = r
        except Exception:
            continue

    # colonnes à ajouter
    wanted = [
        # arbitres
        "arbitre_principal", "nationalite_arbitre",
        "arbitre_assistant1", "nationalite_arbitre_assistant1",
        "arbitre_assistant2", "nationalite_arbitre_assistant2",
        "quatrieme_arbitre", "nationalite_quatrieme_arbitre",
        # DZ +
        "assistant1_dz", "assistant2_dz", "capitaine_dz",
        "carton_rouge_1","carton_rouge_2","carton_rouge_3","carton_rouge_4","carton_rouge_5",
        "carton_jaune_1","carton_jaune_2","carton_jaune_3","carton_jaune_4","carton_jaune_5",
        # ADV
        "entraineur_adv", "assistant1_adv", "assistant2_adv", "capitaine_adv",
    ]
    # ADV players/subs/goals
    wanted += [f"adv_j{i}" for i in range(1, 21)]
    wanted += [f"adv_chang{i}" for i in range(1, 11)]
    wanted += [f"adv_but{i}" for i in range(1, 16)]
    wanted += [f"adv_carton_rouge_{i}" for i in range(1, 6)]
    wanted += [f"adv_carton_jaune_{i}" for i in range(1, 6)]
    # Notes
    wanted += ["notes_qualite"]

    col_index = ensure_columns(ws, header_row, wanted)

    # Détection colonnes existantes DZ: j1..j20, chang1..chang10, but1..but15, entraineur
    # On privilégie les noms de headers déjà présents.
    colmap_norm = build_colmap(ws, header_row)

    def col_by_regex(rx: str) -> Dict[int, int]:
        # retourne mapping n->col
        out = {}
        for k, c in colmap_norm.items():
            m = re.match(rx, k)
            if m:
                out[int(m.group(1))] = c
        return out

    dz_j = col_by_regex(r"^j(\d+)$")
    dz_ch = col_by_regex(r"^chang(\d+)$")
    dz_bu = col_by_regex(r"^but(\d+)$")

    # entraineur dz: souvent "entraineur algérie" ou similaire
    ent_dz_col = None
    for k, c in colmap_norm.items():
        if "entraineur" in k and ("alger" in k or "dz" in k):
            ent_dz_col = c
            break
    # fallback: si une colonne nommée "entraîneur" existe
    if ent_dz_col is None:
        for k, c in colmap_norm.items():
            if k == "entraineur":
                ent_dz_col = c
                break

    stats = {"matched_rows": 0, "filled_green": 0, "improved_orange": 0, "conflicts": 0}
    conflicts_list = []

    for _, row in df_md.iterrows():
        mno = int(row["match_no"])
        if mno not in idx:
            continue
        excel_row = idx[mno]
        stats["matched_rows"] += 1

        # arbitres
        apply_cell(ws, excel_row, col_index["arbitre_principal"], row.get("arbitre_principal",""), stats)
        apply_cell(ws, excel_row, col_index["nationalite_arbitre"], row.get("nationalite_arbitre",""), stats)

        # assistants DZ
        apply_cell(ws, excel_row, col_index["assistant1_dz"], row.get("assistant1_dz",""), stats)
        apply_cell(ws, excel_row, col_index["assistant2_dz"], row.get("assistant2_dz",""), stats)
        apply_cell(ws, excel_row, col_index["capitaine_dz"], row.get("capitaine_dz",""), stats)

        # coach DZ (si colonne détectée)
        if ent_dz_col is not None:
            apply_cell(ws, excel_row, ent_dz_col, row.get("entraineur_dz",""), stats)

        # joueurs DZ : ne remplis que si la colonne existe et la cellule est vide (sinon conflit)
        for i in range(1, 21):
            val = row.get(f"dz_j{i}", "")
            if i in dz_j:
                apply_cell(ws, excel_row, dz_j[i], val, stats)

        # changements DZ
        for i in range(1, 11):
            val = row.get(f"dz_chang{i}", "")
            if i in dz_ch:
                apply_cell(ws, excel_row, dz_ch[i], val, stats)

        # buts DZ
        for i in range(1, 16):
            val = row.get(f"dz_but{i}", "")
            if i in dz_bu:
                apply_cell(ws, excel_row, dz_bu[i], val, stats)

        # ADV colonnes ajoutées
        apply_cell(ws, excel_row, col_index["entraineur_adv"], row.get("entraineur_adv",""), stats)
        apply_cell(ws, excel_row, col_index["assistant1_adv"], row.get("assistant1_adv",""), stats)
        apply_cell(ws, excel_row, col_index["assistant2_adv"], row.get("assistant2_adv",""), stats)
        apply_cell(ws, excel_row, col_index["capitaine_adv"], row.get("capitaine_adv",""), stats)

        for i in range(1, 21):
            apply_cell(ws, excel_row, col_index[f"adv_j{i}"], row.get(f"adv_j{i}",""), stats)
        for i in range(1, 11):
            apply_cell(ws, excel_row, col_index[f"adv_chang{i}"], row.get(f"adv_chang{i}",""), stats)
        for i in range(1, 16):
            apply_cell(ws, excel_row, col_index[f"adv_but{i}"], row.get(f"adv_but{i}",""), stats)

        # notes_qualite: écrit seulement si cellule vide
        note = row.get("notes_global","")
        if note:
            apply_cell(ws, excel_row, col_index["notes_qualite"], note, stats)

    wb.save(out_path)
    return stats


# ------------------------------
# Main
# ------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--md", required=True, help="Path Matches EN 1963 -15 janv.2026.md")
    ap.add_argument("--xlsx", required=True, help="Path en_matches.xlsx")
    ap.add_argument("--outdir", required=True, help="Output directory")
    args = ap.parse_args()

    with open(args.md, "r", encoding="utf-8") as f:
        md_text = f.read()

    md_matches = iter_md_matches(md_text)
    parsed = [parse_one_match(m) for m in md_matches]
    df = pd.DataFrame(parsed)

    # Ordre colonnes MD structured
    base_cols = [
        "match_no","date","team_a","team_b","score_a","score_b",
        "stade","ville","competition_edition_etape",
        "arbitre_principal","nationalite_arbitre",
        "arbitre_assistant1","nationalite_arbitre_assistant1",
        "arbitre_assistant2","nationalite_arbitre_assistant2",
        "quatrieme_arbitre","nationalite_quatrieme_arbitre",
        "entraineur_dz","assistant1_dz","assistant2_dz","capitaine_dz",
        "entraineur_adv","assistant1_adv","assistant2_adv","capitaine_adv",
    ]
    dz_cols = [f"dz_j{i}" for i in range(1,21)] + [f"dz_chang{i}" for i in range(1,11)] + [f"dz_but{i}" for i in range(1,16)]
    adv_cols = [f"adv_j{i}" for i in range(1,21)] + [f"adv_chang{i}" for i in range(1,11)] + [f"adv_but{i}" for i in range(1,16)]
    cards = (
        [f"carton_rouge_{i}" for i in range(1,6)] + [f"carton_jaune_{i}" for i in range(1,6)] +
        [f"adv_carton_rouge_{i}" for i in range(1,6)] + [f"adv_carton_jaune_{i}" for i in range(1,6)]
    )
    cols = base_cols + dz_cols + adv_cols + cards + ["notes_global"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]

    out_md = f"{args.outdir.rstrip('/')}/matches_md_structured.xlsx"
    df.to_excel(out_md, index=False)

    out_xlsx = f"{args.outdir.rstrip('/')}/en_matches_complet.xlsx"
    stats = update_xlsx(args.xlsx, out_xlsx, df)

    print("=== Résumé ===")
    print(f"Matchs parsés MD: {len(df)}")
    print(f"Lignes matchées dans XLSX: {stats['matched_rows']}")
    print(f"Cellules remplies (VERT): {stats['filled_green']}")
    print(f"Cellules améliorées (ORANGE): {stats['improved_orange']}")
    print(f"Conflits non écrasés: {stats['conflicts']}")
    print(f"Sorties: {out_md} | {out_xlsx}")

if __name__ == "__main__":
    main()
