#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
build_adv_players.py

Génère un Excel (et optionnellement un CSV) à partir d'un fichier Markdown de matchs.

Sortie: 1 ligne par match (1..679)
Colonnes:
- N° du match
- Equipe adverse
- entraineur_adv
- entraineur_assistant_adv
- adv_j1..adv_j20 (adv_j1..adv_j11 titulaires, adv_j12..adv_j20 entrants)
- adv_chang1_sortant..adv_chang10_sortant (sortants + minute si présente)
- notes (surplus coachs, anomalies)

Règles clés:
- parenthèses => entrant ; le sortant est le joueur juste avant
- minutes uniquement dans adv_changX_sortant ; jamais dans adv_jX
- joueurs séparés par virgules OU tiret avec espaces (" - ", " – ", " — ")
- ne pas splitter sur la virgule interne d'une parenthèse: (Nom, 61')
- coach adverse: après "Entraineur:"/"Entraineurs:" (avec ou sans espaces),
  parfois sur la même ligne que les joueurs, parfois sur la ligne suivante.
- cas "A (B) C" => A sort remplacé par B, puis C est un autre joueur.

Exécution:
  python3 build_adv_players.py --md ./MatchesEN.md --out ./file.xlsx
Optionnel:
  python3 build_adv_players.py --md ./MatchesEN.md --out ./file.xlsx --csv ./file.csv
"""

import argparse
import csv
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

MAX_MATCH = 679
MAX_PLAYERS = 20
MAX_STARTERS = 11
MAX_CHANGES = 10


def read_text_fallback(path: str) -> str:
    data = Path(path).read_bytes()

    # BOM UTF-16 (FF FE / FE FF)
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        return data.decode("utf-16")

    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue

    return data.decode("latin-1", errors="replace")


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s or "")
        if unicodedata.category(c) != "Mn"
    )

def norm_key(s: str) -> str:
    s = strip_accents((s or "").lower())
    s = s.replace("’", "'")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def clean_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def strip_md(line: str) -> str:
    s = (line or "").strip()
    s = s.replace("**", "").replace("__", "")
    s = re.sub(r"^[#\s>*`_]+", "", s)
    return s.strip()

def strip_trailing_punct(s: str) -> str:
    return clean_spaces(re.sub(r"[,\.;:\-]+$", "", (s or "").strip()))

def remove_captain_markers(text: str) -> str:
    t = text or ""
    t = t.replace("©", " ")
    t = re.sub(r"\((?:cap|c|C|CAP)\)", " ", t)
    t = re.sub(r"\bcap\b", " ", t, flags=re.IGNORECASE)
    return clean_spaces(t)

def clean_player_name(name: str) -> str:
    t = clean_spaces(name)
    t = remove_captain_markers(t)
    t = re.sub(r"^\s*\d{1,2}\s*-\s*", "", t)
    t = re.sub(r"\b\d{1,3}(?:\+\d{1,2})?\s*[’']\s*$", "", t).strip()
    t = strip_trailing_punct(t)
    return t


RE_MATCH_NO = re.compile(r"\(\s*(\d{1,3})\s*\)")
RE_HEADER_SCORE = re.compile(r"^(.*?)\s+(\d+)\s+(.*?)\s+(\d+)\s*$", re.U)

def is_match_header(line: str) -> bool:
    l = strip_md(line)
    return bool(RE_MATCH_NO.search(l)) and bool(re.search(r"\d+\s+\D+\s+\d+", l))

def parse_header_line(line: str) -> Optional[Tuple[int, str, int, str, int]]:
    raw = strip_md(line)
    m = RE_MATCH_NO.search(raw)
    if not m:
        return None
    no = int(m.group(1))

    after = raw[m.end():].strip()
    after = after.replace(",", " ")
    after = re.sub(r"\s+", " ", after).strip()

    m2 = RE_HEADER_SCORE.match(after)
    if not m2:
        return None

    teamA = clean_spaces(m2.group(1))
    scoreA = int(m2.group(2))
    teamB = clean_spaces(m2.group(3))
    scoreB = int(m2.group(4))
    return no, teamA, scoreA, teamB, scoreB

def guess_adversary(teamA: str, teamB: str) -> Optional[str]:
    a = norm_key(teamA)
    b = norm_key(teamB)
    if a in {"algerie", "algeria"} and b not in {"algerie", "algeria"}:
        return teamB
    if b in {"algerie", "algeria"} and a not in {"algerie", "algeria"}:
        return teamA
    return None


def split_team_line(line: str) -> Optional[Tuple[str, str]]:
    l = strip_md(line)
    if ":" not in l:
        return None
    left, right = l.split(":", 1)
    left = clean_spaces(left)
    if re.match(r"^Entraineurs?\b", left, flags=re.I):
        return None
    if len(left) < 2:
        return None
    return left, right.strip()


COACH_SPLIT_RE = re.compile(r"(?:^|[,\.;]\s*|\s+)\bEntraineurs?\b\s*:\s*", re.I)

def split_coach_names(text: str) -> List[str]:
    t = clean_spaces(text).strip().strip(".")
    t = t.replace(";", ",").replace(".", ",")
    t = re.sub(r"\s+et\s+", ",", t, flags=re.I)
    parts = [strip_trailing_punct(clean_spaces(p)) for p in t.split(",")]
    parts = [p for p in parts if p]
    out, seen = [], set()
    for p in parts:
        k = norm_key(p)
        if k and k not in seen:
            out.append(p)
            seen.add(k)
    return out

def extract_coach_from_payload(payload: str) -> Tuple[str, Optional[str], Optional[str], List[str]]:
    s = clean_spaces(payload)
    m = COACH_SPLIT_RE.search(s)
    if not m:
        return s, None, None, []
    players_part = clean_spaces(s[:m.start()]).rstrip(" ,;.")
    coach_part = clean_spaces(s[m.end():])
    names = split_coach_names(coach_part)
    coach = names[0] if len(names) >= 1 else None
    assistant = names[1] if len(names) >= 2 else None
    extras = names[2:] if len(names) > 2 else []
    return players_part, coach, assistant, extras


def split_players_outside_parens(text: str) -> List[str]:
    s = text or ""
    tokens, buf = [], []
    depth = 0
    i = 0

    def flush():
        t = clean_spaces("".join(buf))
        t = strip_trailing_punct(t)
        if t:
            tokens.append(t)
        buf.clear()

    while i < len(s):
        ch = s[i]
        if ch == "(":
            depth += 1
            buf.append(ch); i += 1; continue
        if ch == ")":
            depth = max(0, depth - 1)
            buf.append(ch); i += 1; continue

        if depth == 0 and ch == ",":
            flush(); i += 1; continue

        if depth == 0 and s[i:i+3] in (" - ", " – ", " — "):
            flush(); i += 3; continue

        buf.append(ch)
        i += 1

    flush()
    return tokens


RE_PARENS = re.compile(r"\(([^()]*)\)")

def parse_entrant_and_minute(paren_content: str) -> Tuple[str, Optional[str]]:
    c = clean_spaces(paren_content)
    c_norm = norm_key(c)
    if c_norm in {"c", "cap", "capitaine"}:
        return "", None
    c = remove_captain_markers(c)

    m = re.match(r"^(.*?)(?:,\s*|\s+)?(\d{1,3}(?:\+\d{1,2})?)\s*(?:['’]|e)?\s*$", c)
    if m:
        name = clean_player_name(m.group(1))
        minute = m.group(2)
        return (name, f"{minute}'") if name else ("", None)

    return clean_player_name(c), None

def process_fragment(fragment: str, players: List[str], entrants: List[str], changes_out: List[str], notes: List[str]) -> None:
    frag = clean_spaces(fragment)
    if not frag:
        return

    # "A (B) C" => découper en 2 traitements
    while True:
        m = re.search(r"\)\s+(.+)$", frag)
        if not m:
            break
        prefix = frag[:m.start()+1]
        suffix = frag[m.start()+2:].strip()
        process_fragment(prefix, players, entrants, changes_out, notes)
        frag = suffix
        if not frag:
            return

    parens = RE_PARENS.findall(frag)
    base = RE_PARENS.sub(" ", frag)
    base = clean_player_name(base)

    sub_groups = []
    for p in parens:
        entrant, minute = parse_entrant_and_minute(p)
        if entrant:
            sub_groups.append((entrant, minute))

    if base:
        players.append(base)
        for entrant, minute in sub_groups:
            if len(changes_out) < MAX_CHANGES:
                changes_out.append(f"{base}{' ' + minute if minute else ''}".strip())
            else:
                notes.append(f"Changements>10: sortant={base}")
            if len(entrants) < (MAX_PLAYERS - MAX_STARTERS):
                entrants.append(entrant)
            else:
                notes.append(f"Entrants>9: entrant={entrant}")
        return

    if sub_groups:
        entrant, minute = sub_groups[0]
        if players:
            sortant = players[-1]
            if len(changes_out) < MAX_CHANGES:
                changes_out.append(f"{sortant}{' ' + minute if minute else ''}".strip())
            else:
                notes.append(f"Changements>10: sortant={sortant}")
        else:
            notes.append(f"Entrant sans sortant: {entrant}")

        if len(entrants) < (MAX_PLAYERS - MAX_STARTERS):
            entrants.append(entrant)
        else:
            notes.append(f"Entrants>9: entrant={entrant}")

        if len(sub_groups) > 1:
            notes.append(f"Multi-parenthèses entrant: {sub_groups[1:]}")
        return


def collect_adv_section(block_lines: List[str], adv_team: str) -> List[str]:
    tnorm = norm_key(adv_team)
    out = []
    started = False

    for raw in block_lines:
        l = strip_md(raw)
        if not l:
            continue

        if started and is_match_header(l):
            break

        tl = split_team_line(l)
        if tl:
            team, payload = tl
            if norm_key(team) == tnorm:
                started = True
                out.append(payload)
                continue
            if started and norm_key(team) != tnorm:
                break

        if started:
            out.append(l)

    return out


def parse_md(md_text: str) -> Dict[int, dict]:
    lines = [strip_md(l) for l in md_text.splitlines()]

    headers = []
    for i, line in enumerate(lines):
        if is_match_header(line):
            h = parse_header_line(line)
            if h and 1 <= h[0] <= MAX_MATCH:
                headers.append((i, h))
    headers.sort(key=lambda x: x[0])

    by_no = {}
    for idx, h in headers:
        if h[0] not in by_no:
            by_no[h[0]] = (idx, h)

    results = {}
    for no in range(1, MAX_MATCH + 1):
        results[no] = {
            "N° du match": no,
            "Equipe adverse": None,
            "entraineur_adv": None,
            "entraineur_assistant_adv": None,
            "adv_players": [],
            "adv_entrants": [],
            "adv_changes_out": [],
            "notes": [],
        }

        if no not in by_no:
            continue

        start_idx, header = by_no[no]
        _, teamA, _, teamB, _ = header
        adv = guess_adversary(teamA, teamB)
        if not adv:
            results[no]["notes"].append("Adversaire non détecté (Algérie absente de l'entête)")
            continue
        results[no]["Equipe adverse"] = adv

        # bloc match
        next_start = None
        for idx, _h in headers:
            if idx > start_idx:
                next_start = idx
                break
        end_idx = next_start if next_start is not None else len(lines)
        block = lines[start_idx:end_idx]

        adv_section = collect_adv_section(block, adv)
        if not adv_section:
            results[no]["notes"].append("Section adverse introuvable (ligne équipe absente)")
            continue

        section_join = " ".join([clean_spaces(x) for x in adv_section if clean_spaces(x)])
        section_join = re.sub(r"\s+", " ", section_join).strip()

        players_text, coach, assistant, extras = extract_coach_from_payload(section_join)
        if coach:
            results[no]["entraineur_adv"] = coach
        if assistant:
            results[no]["entraineur_assistant_adv"] = assistant
        if extras:
            results[no]["notes"].append(f"Coachs supplémentaires: {extras}")

        tokens = split_players_outside_parens(players_text)

        players, entrants, changes_out = [], [], []
        notes = results[no]["notes"]

        for tok in tokens:
            tok = clean_spaces(tok)
            if tok:
                process_fragment(tok, players, entrants, changes_out, notes)

        # dédup joueurs/entrants (ordre conservé)
        def dedup(seq):
            out, seen = [], set()
            for x in seq:
                x = clean_player_name(x)
                if not x:
                    continue
                k = norm_key(x)
                if k and k not in seen:
                    out.append(x); seen.add(k)
            return out

        players = dedup(players)
        entrants = dedup(entrants)

        # nettoyage changements (minutes autorisées)
        cleaned = []
        for c in changes_out:
            c = strip_trailing_punct(clean_spaces(c))
            if c:
                cleaned.append(c)
        changes_out = cleaned

        if len(players) > MAX_PLAYERS:
            notes.append(f"Joueurs>20: {len(players)} (tronqué)")
            players = players[:MAX_PLAYERS]
        if len(entrants) > (MAX_PLAYERS - MAX_STARTERS):
            notes.append(f"Entrants>9: {len(entrants)} (tronqué)")
            entrants = entrants[:(MAX_PLAYERS - MAX_STARTERS)]
        if len(changes_out) > MAX_CHANGES:
            notes.append(f"Changements>10: {len(changes_out)} (tronqué)")
            changes_out = changes_out[:MAX_CHANGES]

        results[no]["adv_players"] = players
        results[no]["adv_entrants"] = entrants
        results[no]["adv_changes_out"] = changes_out

    return results


def build_columns() -> List[str]:
    cols = ["N° du match", "Equipe adverse", "entraineur_adv", "entraineur_assistant_adv"]
    cols += [f"adv_j{i}" for i in range(1, MAX_PLAYERS + 1)]
    cols += [f"adv_chang{i}_sortant" for i in range(1, MAX_CHANGES + 1)]
    cols += ["notes"]
    return cols

def write_xlsx(rows: Dict[int, dict], out_path: str) -> None:
    cols = build_columns()
    wb = Workbook()
    ws = wb.active
    ws.title = "matches_adv"
    ws.append(cols)

    for no in range(1, MAX_MATCH + 1):
        r = rows[no]

        players_all = (r["adv_players"] or [])
        entrants = (r["adv_entrants"] or [])

        starters = players_all[:MAX_STARTERS]
        listed_after_11 = players_all[MAX_STARTERS:]

        bench_in = entrants + listed_after_11
        bench_in = bench_in[:(MAX_PLAYERS - MAX_STARTERS)]

        if listed_after_11 and not entrants:
            r["notes"].append(
                f"{len(listed_after_11)} joueur(s) listé(s) après les 11 sans remplacements explicites"
            )

        adv_j = starters + bench_in
        adv_j += [None] * (MAX_PLAYERS - len(adv_j))

        chang = (r["adv_changes_out"] or [])[:MAX_CHANGES]
        chang += [None] * (MAX_CHANGES - len(chang))

        notes = "".join(r["notes"]) if r["notes"] else None

        row = [r["N° du match"], r["Equipe adverse"], r["entraineur_adv"], r["entraineur_assistant_adv"]]
        row += adv_j
        row += chang
        row += [notes]
        ws.append(row)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    for i in range(5, 5 + MAX_PLAYERS + MAX_CHANGES):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions[get_column_letter(5 + MAX_PLAYERS + MAX_CHANGES)].width = 50

    wb.save(out_path)

def write_csv(rows: Dict[int, dict], csv_path: str) -> None:
    cols = build_columns()
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for no in range(1, MAX_MATCH + 1):
            r = rows[no]

            players_all = (r["adv_players"] or [])
            entrants = (r["adv_entrants"] or [])

            starters = players_all[:MAX_STARTERS]
            listed_after_11 = players_all[MAX_STARTERS:]

            bench_in = entrants + listed_after_11
            bench_in = bench_in[:(MAX_PLAYERS - MAX_STARTERS)]

            if listed_after_11 and not entrants:
                r["notes"].append(
                    f"{len(listed_after_11)} joueur(s) listé(s) après les 11 sans remplacements explicites"
                )

            adv_j = starters + bench_in
            adv_j += [""] * (MAX_PLAYERS - len(adv_j))

            chang = (r["adv_changes_out"] or [])[:MAX_CHANGES]
            chang += [""] * (MAX_CHANGES - len(chang))

            notes = "".join(r["notes"]) if r["notes"] else ""
            row = [
                r["N° du match"],
                r["Equipe adverse"] or "",
                r["entraineur_adv"] or "",
                r["entraineur_assistant_adv"] or "",
            ]
            row += adv_j
            row += chang
            row += [notes]
            w.writerow(row)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--md", required=True, help="Chemin du Markdown")
    ap.add_argument("--out", required=True, help="Chemin du XLSX")
    ap.add_argument("--csv", required=False, help="Optionnel: chemin CSV")
    args = ap.parse_args()

    md_text = read_text_fallback(args.md)
    rows = parse_md(md_text)

    write_xlsx(rows, args.out)
    if args.csv:
        write_csv(rows, args.csv)

    parsed = sum(1 for i in range(1, MAX_MATCH + 1) if rows[i]["Equipe adverse"])
    print(f"[OK] {args.out}")
    print(f"[INFO] Matchs détectés (adversaire trouvé): {parsed}/{MAX_MATCH}")

if __name__ == "__main__":
    main()
