#!/usr/bin/env python3

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


KNOWN_PREFIXES = (
    "ne le",
    "né le",
    "née le",
    "décédé",
    "decede",
    "carrière sportive",
    "carriere sportive",
    "les matches",
    "le matches",
    "nombre de matches",
    "nombre de victoire",
    "nombre de défaite",
    "nombre de defaite",
    "nombre de nuls",
    "- nombre de",
)

MONTHS = {
    "janvier": 1,
    "fevrier": 2,
    "fevrier.": 2,
    "fevrier,": 2,
    "fevrier)": 2,
    "février": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "ami": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "août": 8,
    "septembre": 9,
    "octobre": 10,
    "novembre": 11,
    "decembre": 12,
    "décembre": 12,
}

COUNTRIES = {
    "algerie": "Algérie",
    "algérie": "Algérie",
    "france": "France",
    "belgique": "Belgique",
    "roumanie": "Roumanie",
    "maroc": "Maroc",
    "espagne": "Espagne",
    "bosnie-herzegovine": "Bosnie-Herzégovine",
    "bosnie-herzégovine": "Bosnie-Herzégovine",
    "yougoslavie": "Yougoslavie",
}

STAT_PATTERNS = [
    (
        re.compile(r"nombre de s[ée]lections comme joueur en equipe de france", re.I),
        "nombre_selections_joueur_equipe_france",
    ),
    (
        re.compile(r"nombre de s[ée]lections comme joueur", re.I),
        "nombre_selections_joueur",
    ),
    (
        re.compile(r"nombre de matches comme entr[aâ][iî]neur principal", re.I),
        "nombre_matches_entraineur_principal",
    ),
    (
        re.compile(r"nombre de matches comme entr[aâ][iî]neur national adjoint", re.I),
        "nombre_matches_entraineur_national_adjoint",
    ),
    (
        re.compile(r"nombre de matches comme co[- ]?entr[aâ][iî]neur national", re.I),
        "nombre_matches_co_entraineur_national",
    ),
    (
        re.compile(r"nombre de matches comme entr[aâ][iî]neur national", re.I),
        "nombre_matches_entraineur_national",
    ),
]

SUMMARY_PATTERNS = [
    (re.compile(r"^nombre de matches\b", re.I), "bilan_nombre_matches"),
    (re.compile(r"^nombre de victoire\b", re.I), "bilan_nombre_victoires"),
    (re.compile(r"^nombre de défaite\b", re.I), "bilan_nombre_defaites"),
    (re.compile(r"^nombre de defaite\b", re.I), "bilan_nombre_defaites"),
    (re.compile(r"^nombre de nuls\b", re.I), "bilan_nombre_nuls"),
]


def clean_line(text: str) -> str:
    text = text.replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def ascii_fold(text: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


def normalize_key(text: str) -> str:
    return ascii_fold(text).lower().strip()


def canonical_country(text: str) -> str:
    key = normalize_key(text.strip(" ,.;:()[]"))
    return COUNTRIES.get(key, text.strip(" ,.;:()[]"))


def is_country(text: str) -> bool:
    return normalize_key(text.strip(" ,.;:()[]")) in COUNTRIES


def is_placeholder_line(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    punctuation = stripped.replace(" ", "")
    if set(punctuation) <= {".", "-", "_"}:
        return True
    return bool(re.fullmatch(r"[.\-_]{3,}", punctuation))


def numeric_date_to_ddmmyyyy(text: str) -> str:
    match = re.match(r"(\d{1,2})[./](\d{1,2})[./](\d{2,4})", text.strip())
    if not match:
        return ""
    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    if year < 100:
        year += 1900
    if year < 1000 or not 1 <= day <= 31 or not 1 <= month <= 12:
        return ""
    return f"{day:02d}/{month:02d}/{year:04d}"


def french_date_to_ddmmyyyy(text: str) -> str:
    match = re.match(r"(\d{1,2})\s+([A-Za-zÀ-ÿ-]+)\s+(\d{4})", text.strip())
    if not match:
        return ""
    day = int(match.group(1))
    month_name = normalize_key(match.group(2).replace(".", ""))
    month = MONTHS.get(month_name)
    if not month:
        return ""
    year = int(match.group(3))
    return f"{day:02d}/{month:02d}/{year:04d}"


def extract_date_prefix(text: str) -> tuple[str, str, str]:
    cleaned = text.strip(" ,.;:")
    full_patterns = [
        re.compile(r"^(\d{1,2}[./]\d{1,2}[./]\d{2,4})"),
        re.compile(r"^(\d{1,2}\s+[A-Za-zÀ-ÿ-]+\s+\d{4})"),
    ]
    for pattern in full_patterns:
        match = pattern.match(cleaned)
        if match:
            date_source = match.group(1).strip()
            normalized = numeric_date_to_ddmmyyyy(date_source) or french_date_to_ddmmyyyy(date_source)
            remainder = cleaned[match.end() :].strip(" ,.;:")
            return normalized, date_source, remainder

    partial_match = re.match(r"^([A-Za-zÀ-ÿ-]+\s+\d{4})", cleaned)
    if partial_match:
        date_source = partial_match.group(1).strip()
        remainder = cleaned[partial_match.end() :].strip(" ,.;:")
        return "", date_source, remainder

    return "", "", cleaned


def strip_location_lead(text: str) -> str:
    text = text.strip(" ,.;:")
    text = re.sub(r"\s+[aà] l['’]âge.*$", "", text, flags=re.I)
    text = re.sub(r"\s+[aà] l['’]age.*$", "", text, flags=re.I)
    text = text.strip(" ,.;:[]")
    text = re.sub(r"^(?:,?\s*)?(?:à|a|au|en)\s*", "", text, flags=re.I)
    return text.strip(" ,.;:[]")


def split_parenthetical(text: str) -> tuple[str, str]:
    match = re.match(r"^(.*?)\(([^()]*)\)?$", text)
    if not match:
        return text.strip(), ""
    before = match.group(1).strip(" ,")
    inside = match.group(2).strip(" ,")
    return before, inside


def parse_place(text: str) -> dict[str, str]:
    raw = strip_location_lead(text)
    result = {
        "source": raw,
        "ville": "",
        "region": "",
        "pays": "",
    }
    if not raw:
        return result

    base, inside = split_parenthetical(raw)
    if inside:
        parts = [part.strip(" ,") for part in inside.split("-") if part.strip(" ,")]
        result["ville"] = base
        if len(parts) == 1:
            if is_country(parts[0]):
                result["pays"] = canonical_country(parts[0])
            else:
                result["region"] = parts[0]
        elif len(parts) >= 2:
            result["region"] = parts[0]
            if is_country(parts[1]):
                result["pays"] = canonical_country(parts[1])
            else:
                result["pays"] = parts[1]
        return result

    if "," in raw:
        parts = [part.strip(" ,") for part in raw.split(",") if part.strip(" ,")]
        if parts:
            if len(parts) == 1:
                if is_country(parts[0]):
                    result["pays"] = canonical_country(parts[0])
                else:
                    result["ville"] = parts[0]
            elif len(parts) == 2:
                result["ville"] = parts[0]
                normalized_part = strip_location_lead(parts[1]).strip(" -")
                if is_country(normalized_part):
                    result["pays"] = canonical_country(normalized_part)
                else:
                    result["region"] = parts[1]
            else:
                result["ville"] = parts[0]
                result["region"] = parts[1]
                normalized_part = strip_location_lead(parts[2]).strip(" -")
                result["pays"] = canonical_country(normalized_part) if is_country(normalized_part) else normalized_part
            return result

    if re.search(r"\ben\b", raw, re.I):
        left, right = re.split(r"\ben\b", raw, maxsplit=1, flags=re.I)
        left = left.strip(" ,")
        right = right.strip(" ,")
        if left:
            if is_country(left):
                result["pays"] = canonical_country(left)
            else:
                result["ville"] = left
        if right:
            if is_country(right):
                result["pays"] = canonical_country(right)
            elif not result["region"]:
                result["region"] = right
        return result

    if re.search(r"\s+à\s+", raw, re.I):
        left, right = re.split(r"\s+à\s+", raw, maxsplit=1, flags=re.I)
        left = left.strip(" ,")
        right = right.strip(" ,")
        result["ville"] = left
        if is_country(right):
            result["pays"] = canonical_country(right)
        else:
            result["region"] = right
        return result

    if is_country(raw):
        result["pays"] = canonical_country(raw)
        return result

    result["ville"] = raw.strip(" .")
    return result


def split_event_details(value: str) -> dict[str, str]:
    normalized_date, date_source, remainder = extract_date_prefix(value)
    place = parse_place(remainder)
    return {
        "date": normalized_date,
        "date_source": date_source,
        "lieu_source": place["source"],
        "ville": place["ville"],
        "region": place["region"],
        "pays": place["pays"],
    }


def is_match_line(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    if is_placeholder_line(stripped):
        return True
    if re.match(r"^\d{1,2}[./]\d{1,2}[./]\d{2,4}", stripped):
        return True
    if re.match(r"^\d+/\d+\s+\d{1,2}[./]\d{1,2}[./]\d{2,4}", stripped):
        return True
    if re.match(r"^[.\-]*\d{1,2}\s+[A-Za-zÀ-ÿ]+\s+\d{4}", stripped):
        return True
    return False


def looks_like_name(text: str) -> bool:
    stripped = text.strip()
    if not stripped or len(stripped) > 90:
        return False
    lowered = stripped.lower()
    if any(lowered.startswith(prefix) for prefix in KNOWN_PREFIXES):
        return False
    if ":" in stripped:
        return False
    if stripped.endswith("."):
        return False
    if len(re.findall(r"[A-Za-zÀ-ÿ]", stripped)) < 4:
        return False
    if re.search(r"\d", stripped):
        return False
    words = re.findall(r"[A-Za-zÀ-ÿ'’]+", stripped)
    if not 2 <= len(words) <= 8:
        return False
    if stripped == stripped.upper():
        return True
    return sum(1 for word in words if word[:1].isupper()) >= max(2, len(words) - 1)


def candidate_name(lines: list[str], index: int) -> bool:
    text = lines[index]
    if not looks_like_name(text):
        return False
    if index == len(lines) - 1:
        return True

    indicators = 0
    for line in lines[index + 1 : index + 5]:
        lowered = line.lower()
        if lowered.startswith(("né", "ne ", "décédé", "decede", "- nombre de", "nombre de")):
            indicators += 1
        if "carrière sportive" in lowered or "carriere sportive" in lowered:
            indicators += 1
        if lowered.startswith(("les matches", "le matches")):
            indicators += 1
    if indicators >= 1:
        return True

    next_line = lines[index + 1] if index + 1 < len(lines) else ""
    return bool(next_line and len(next_line.split()) >= 12 and not looks_like_name(next_line))


def extract_value(text: str) -> str:
    text = re.sub(r"^[\-–]\s*", "", text).strip()
    text = re.sub(r"^(?:né|née|décédé|decede)\s*(?:le\s*)?:?\s*", "", text, flags=re.I)
    if ":" in text:
        return text.split(":", 1)[1].strip()
    return text


def parse_integer(text: str) -> int | None:
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else None


def parse_match_details(line: str) -> dict[str, str]:
    if is_placeholder_line(line):
        return {
            "type_ligne": "placeholder",
            "date_match": "",
            "date_match_source": "",
            "detail_match_source": "",
        }

    work = line.strip()
    score_prefix = ""
    leading_score = re.match(r"^(\d+/\d+)\s+(.*)$", work)
    if leading_score:
        score_prefix = leading_score.group(1)
        work = leading_score.group(2).strip()

    date_patterns = [
        re.compile(r"^(\d{1,2}[./]\d{1,2}[./]\d{2,4})(.*)$"),
        re.compile(r"^([.\-]*\d{1,2}\s+[A-Za-zÀ-ÿ]+\s+\d{4})(.*)$"),
    ]
    date_match = ""
    date_source = ""
    detail = work
    for pattern in date_patterns:
        match = pattern.match(work)
        if match:
            date_source = match.group(1).strip(" .-")
            date_match = numeric_date_to_ddmmyyyy(date_source) or french_date_to_ddmmyyyy(date_source)
            detail = match.group(2).strip(" :")
            break

    if score_prefix:
        detail = f"{score_prefix} {detail}".strip()

    return {
        "type_ligne": "match",
        "date_match": date_match,
        "date_match_source": date_source,
        "detail_match_source": detail,
    }


def empty_record(name: str) -> dict:
    return {
        "nom_complet": name,
        "date_naissance": "",
        "date_naissance_source": "",
        "lieu_naissance_source": "",
        "ville_naissance": "",
        "region_naissance": "",
        "pays_naissance": "",
        "date_deces": "",
        "date_deces_source": "",
        "lieu_deces_source": "",
        "ville_deces": "",
        "region_deces": "",
        "pays_deces": "",
        "carriere_sportive": "",
        "essentiel_carriere": "",
        "nombre_selections_joueur": None,
        "nombre_selections_joueur_equipe_france": None,
        "nombre_matches_entraineur_national": None,
        "nombre_matches_co_entraineur_national": None,
        "nombre_matches_entraineur_principal": None,
        "nombre_matches_entraineur_national_adjoint": None,
        "bilan_nombre_matches": None,
        "bilan_nombre_victoires": None,
        "bilan_nombre_defaites": None,
        "bilan_nombre_nuls": None,
        "description": "",
        "autres_infos": "",
        "bloc_source": "",
    }


def parse_document(docx_path: Path) -> tuple[list[dict], list[dict]]:
    doc = Document(str(docx_path))
    lines: list[str] = []
    for paragraph in doc.paragraphs:
        for fragment in paragraph.text.splitlines():
            cleaned = clean_line(fragment)
            if cleaned:
                lines.append(cleaned)

    starts = [idx for idx in range(len(lines)) if candidate_name(lines, idx)]
    records: list[dict] = []
    matches_sheet_rows: list[dict] = []

    for pos, start in enumerate(starts):
        end = starts[pos + 1] if pos + 1 < len(starts) else len(lines)
        block = lines[start:end]
        record = empty_record(block[0])
        record["bloc_source"] = "\n".join(block)

        match_lines: list[str] = []
        description_lines: list[str] = []
        extra_lines: list[str] = []
        in_matches = False
        in_description = False

        for line in block[1:]:
            lowered = line.lower()
            stripped_line = re.sub(r"^[\-–]\s*", "", line).strip()
            stripped_lowered = stripped_line.lower()

            if lowered.startswith(("les matches", "le matches")):
                in_matches = True
                in_description = False
                continue

            if stripped_lowered.startswith(("né", "ne ", "née")) and not record["date_naissance_source"] and not record["date_naissance"]:
                details = split_event_details(extract_value(line))
                record["date_naissance"] = details["date"]
                record["date_naissance_source"] = details["date_source"]
                record["lieu_naissance_source"] = details["lieu_source"]
                record["ville_naissance"] = details["ville"]
                record["region_naissance"] = details["region"]
                record["pays_naissance"] = details["pays"]
                continue

            if stripped_lowered.startswith(("décédé", "decede")) and not record["date_deces_source"] and not record["date_deces"]:
                details = split_event_details(extract_value(line))
                record["date_deces"] = details["date"]
                record["date_deces_source"] = details["date_source"]
                record["lieu_deces_source"] = details["lieu_source"]
                record["ville_deces"] = details["ville"]
                record["region_deces"] = details["region"]
                record["pays_deces"] = details["pays"]
                continue

            if "carrière sportive" in lowered or "carriere sportive" in lowered:
                record["carriere_sportive"] = extract_value(line)
                continue

            if "essentiel de la carrière" in lowered or "essentiel de la carriere" in lowered:
                record["essentiel_carriere"] = extract_value(line)
                continue

            matched_stat = False
            for pattern, key in STAT_PATTERNS:
                if pattern.search(lowered):
                    record[key] = parse_integer(line)
                    matched_stat = True
                    break
            if matched_stat:
                continue

            matched_summary = False
            for pattern, key in SUMMARY_PATTERNS:
                if pattern.search(line):
                    record[key] = parse_integer(line)
                    matched_summary = True
                    break
            if matched_summary:
                continue

            if (in_matches or (not in_description and is_match_line(line))) and is_match_line(line):
                in_matches = True
                match_lines.append(line)
                continue

            if in_matches and not is_match_line(line):
                in_matches = False
                in_description = True

            if in_description:
                description_lines.append(line)
            else:
                extra_lines.append(line)

        if not description_lines and extra_lines:
            description_lines = extra_lines
            extra_lines = []

        record["description"] = "\n".join(description_lines)
        record["autres_infos"] = "\n".join(extra_lines)
        records.append(record)

        for order, match_line in enumerate(match_lines, start=1):
            details = parse_match_details(match_line)
            matches_sheet_rows.append(
                {
                    "entraineur": record["nom_complet"],
                    "ordre": order,
                    "type_ligne": details["type_ligne"],
                    "date_match": details["date_match"],
                    "date_match_source": details["date_match_source"],
                    "detail_match_source": details["detail_match_source"],
                    "match_brut": match_line,
                    "source_doc": str(docx_path),
                }
            )

    return records, matches_sheet_rows


def autosize_worksheet(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 80)
    for idx, width in widths.items():
        worksheet.column_dimensions[get_column_letter(idx)].width = width


def style_sheet(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    worksheet.freeze_panes = "A2"
    autosize_worksheet(worksheet)


def write_workbook(records: list[dict], match_rows: list[dict], output_path: Path, source_doc: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "entraineurs"

    headers = [
        "nom_complet",
        "date_naissance",
        "date_naissance_source",
        "ville_naissance",
        "region_naissance",
        "pays_naissance",
        "lieu_naissance_source",
        "date_deces",
        "date_deces_source",
        "ville_deces",
        "region_deces",
        "pays_deces",
        "lieu_deces_source",
        "carriere_sportive",
        "essentiel_carriere",
        "nombre_selections_joueur",
        "nombre_selections_joueur_equipe_france",
        "nombre_matches_entraineur_national",
        "nombre_matches_co_entraineur_national",
        "nombre_matches_entraineur_principal",
        "nombre_matches_entraineur_national_adjoint",
        "bilan_nombre_matches",
        "bilan_nombre_victoires",
        "bilan_nombre_defaites",
        "bilan_nombre_nuls",
        "description",
        "autres_infos",
        "bloc_source",
        "source_doc",
    ]

    ws.append(headers)
    for record in records:
        row = [record.get(header, "") for header in headers[:-1]]
        row.append(str(source_doc))
        ws.append(row)

    matches_ws = wb.create_sheet("matches")
    match_headers = [
        "entraineur",
        "ordre",
        "type_ligne",
        "date_match",
        "date_match_source",
        "detail_match_source",
        "match_brut",
        "source_doc",
    ]
    matches_ws.append(match_headers)
    for row in match_rows:
        matches_ws.append([row.get(header, "") for header in match_headers])

    style_sheet(ws)
    style_sheet(matches_ws)

    for column in ("Z", "AA", "AB"):
        ws.column_dimensions[column].width = 60
    for column in ("F", "G"):
        matches_ws.column_dimensions[column].width = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("docx_path", type=Path)
    parser.add_argument("output_path", type=Path)
    args = parser.parse_args()

    records, match_rows = parse_document(args.docx_path)
    write_workbook(records, match_rows, args.output_path, args.docx_path)
    print(f"records={len(records)}")
    print(f"matches={len(match_rows)}")
    print(f"output={args.output_path}")


if __name__ == "__main__":
    main()
